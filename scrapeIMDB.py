from bs4 import BeautifulSoup
import requests
import openpyxl


excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title='Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank', 'Movie Name', 'Year Of Release', 'IMDB Rating'])



try:
    source=requests.get('https://imdb.com/chart/top/')
    source.raise_for_status()
    
    soup=BeautifulSoup(source.text,'html.parser')
   
    # with open('x.html','w',encoding='utf-8')as f:
    # f.write(str(html))

    #print(soup.prettify)
    movies=soup.find('tbody',class_="lister-list").find_all('tr')
    # print(len(movies))

    for movie in movies:
        
        name=movie.find('td',class_="titleColumn").a.text

        rank=movie.find('td',class_="titleColumn").get_text(strip=True).split('.')[0]

        year=movie.find('td',class_="titleColumn").span.text.strip('()')

        rating=movie.find('td',class_="ratingColumn imdbRating").strong.text
       
        print(rank, name, year, rating)
        sheet.append([rank, name, year, rating])

except Exception as e:
    print(e)

excel.save('IMDB Movie Ratings.xlsx')